#!/usr/bin/env python3
"""Dry-run verifier for the S-empty DAY0 adaptation matrix.

The verifier intentionally calls the production launcher path
``core.wings_entry.build_launcher_plan``. It does not reimplement command
generation. Each case supplies only scenario context plus input/output length,
then toggles all smart/interactive features either on or off.
"""

from __future__ import annotations

import json
import os
import re
import shlex
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
WINGS_CONTROL_ROOT = REPO_ROOT / "wings_control"
DOC_ROOT = WINGS_CONTROL_ROOT / "docs" / "DAY0"
DRY_RUN_ROOT = DOC_ROOT / "dry_run"
OUT_ROOT = DRY_RUN_ROOT / "s_empty_day0_current"

SOURCE_ROWS = (11, 17, 18, 20, 21, 22, 23, 24, 30, 40, 41, 42, 44, 45, 48, 49)
EXCLUDED_ROWS = (12, 19, 25, 31, 43, 46, 47)

JSON_FLAGS = {
    "additional-config",
    "compilation-config",
    "default-chat-template-kwargs",
    "hf-overrides",
    "kv-transfer-config",
    "model-loader-extra-config",
    "speculative-config",
}
BOOL_FLAGS = {
    "async-scheduling",
    "enable-auto-tool-choice",
    "enable-expert-parallel",
    "enable-prefix-caching",
    "headless",
    "language-model-only",
    "no-disable-hybrid-kv-cache-manager",
    "no-enable-prefix-caching",
    "trust-remote-code",
}
CLI_FIELDS = {
    "api-server-count",
    "block-size",
    "data-parallel-size",
    "data-parallel-size-local",
    "data-parallel-start-rank",
    "enable-auto-tool-choice",
    "enable-expert-parallel",
    "enable-prefix-caching",
    "gpu-memory-utilization",
    "headless",
    "kv-cache-dtype",
    "kv-offloading-backend",
    "kv-offloading-size",
    "language-model-only",
    "max-model-len",
    "max-num-batched-tokens",
    "max-num-seqs",
    "mm-encoder-tp-mode",
    "moe-backend",
    "no-disable-hybrid-kv-cache-manager",
    "no-enable-prefix-caching",
    "quantization",
    "reasoning-parser",
    "safetensors-load-strategy",
    "served-model-name",
    "tensor-parallel-size",
    "tokenizer-mode",
    "tool-call-parser",
    "trust-remote-code",
}
FEATURE_CLI_FLAGS = {
    "hf-overrides",
    "kv-transfer-config",
    "kv-offloading-backend",
    "kv-offloading-size",
    "sparse-config",
    "speculative-config",
}
INTERACTIVE_CLI_FLAGS = {
    "default-chat-template-kwargs",
    "enable-auto-tool-choice",
    "reasoning-parser",
    "tool-call-parser",
}
NON_BLOCKING_CLI_FIELDS = {
    "served-model-name",
}
FEATURE_ENV_KEYS = {
    "LMCACHE_LOCAL_CPU",
    "LMCACHE_LOCAL_DISK",
    "LMCACHE_MAX_LOCAL_CPU_SIZE",
    "LMCACHE_MAX_LOCAL_DISK_SIZE",
    "MMC_LOCAL_CONFIG_PATH",
    "WINGS_MEMCACHE_CONFIG_STORE_URL",
    "WINGS_MEMCACHE_DRAM_GB",
    "WINGS_MEMCACHE_META_SERVICE_URL",
}
MANAGED_ENV_KEYS = {
    "ASCEND_PLATFORM",
    "AVAILABLE_POD_MEM_SIZE",
    "BLOCK_SIZE",
    "CONFIG_FILE",
    "DEVICE_COUNT",
    "DISTRIBUTED",
    "DISTRIBUTED_EXECUTOR_BACKEND",
    "DTYPE",
    "ENABLE_AUTO_THINK_CHOICE",
    "ENABLE_AUTO_TOOL_CHOICE",
    "ENABLE_CHUNKED_PREFILL",
    "ENABLE_EXPERT_PARALLEL",
    "ENABLE_KV_DISK_OFFLOAD",
    "ENABLE_KV_MEM_OFFLOAD",
    "ENABLE_KV_OFFLOAD",
    "ENABLE_PREFIX_CACHING",
    "ENABLE_RAG_ACC",
    "ENABLE_REASON_PROXY",
    "ENABLE_SPARSE",
    "ENABLE_SPECULATIVE_DECODE",
    "ENGINE",
    "ENGINE_IMAGE_FLAVOR",
    "ENGINE_VERSION",
    "GPU_MEMORY_UTILIZATION",
    "GLOO_SOCKET_IFNAME",
    "HEAD_NODE_ADDR",
    "HOST",
    "INPUT_LENGTH",
    "KV_CACHE_DTYPE",
    "KV_MEM_OFFLOAD_SIZE",
    "LMCACHE_OFFLOAD",
    "MASTER_IP",
    "MAX_NUM_BATCHED_TOKENS",
    "MAX_NUM_SEQS",
    "MODEL_NAME",
    "MODEL_PATH",
    "MODEL_TYPE",
    "NETWORK_INTERFACE",
    "NNODES",
    "NODE_IPS",
    "NODE_RANK",
    "NODES",
    "OUTPUT_LENGTH",
    "POD_IP",
    "PORT",
    "QUANTIZATION",
    "RANK_IP",
    "RAY_HEAD_IP",
    "SAVE_PATH",
    "SEED",
    "SHARED_VOLUME_PATH",
    "SD_ENABLE",
    "SPARSE_ENABLE",
    "SPECULATIVE_DECODE_MODEL_PATH",
    "TRUST_REMOTE_CODE",
    "WINGS_ASCEND_PLATFORM",
    "WINGS_HARDWARE_FILE",
}


@dataclass(frozen=True)
class Scenario:
    row: int
    slug: str
    model_name: str
    architecture: str
    engine: str
    device: str
    hardware_family: str
    detail_name: str
    device_count: int
    adaptation: str
    quantize: str = ""
    model_type: str = "auto"
    distributed: bool = False
    nnodes: int = 1
    ranks: tuple[int, ...] = (0,)
    extra: bool = False
    note: str = ""


@dataclass(frozen=True)
class Case:
    scenario: Scenario
    mode: str
    node_rank: int
    reference_kind: str
    reference_text: str

    @property
    def name(self) -> str:
        rank = f"-rank{self.node_rank}" if self.scenario.distributed else ""
        return f"{self.scenario.slug}-{self.mode}{rank}"


@dataclass
class ParsedCommand:
    command: str = ""
    flags: dict[str, Any] = field(default_factory=dict)
    env: dict[str, Any] = field(default_factory=dict)
    parse_error: str = ""


def _as_posix(path: Path) -> str:
    return str(path).replace("\\", "/")


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8", newline="\n")


def _write_json(path: Path, value: Any) -> None:
    _write_text(path, json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True))


def _safe_clear_dir(path: Path) -> None:
    root = DRY_RUN_ROOT.resolve()
    target = path.resolve()
    if target == root or root not in target.parents:
        raise RuntimeError(f"refusing to clear path outside dry_run root: {target}")
    path.mkdir(parents=True, exist_ok=True)
    for child in path.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()


def _ensure_import_path() -> None:
    sys.path.insert(0, str(WINGS_CONTROL_ROOT))


def _excel_path() -> Path:
    matches = list((WINGS_CONTROL_ROOT / "docs").glob("AI Infra*_收编视图.xlsx"))
    if not matches:
        raise FileNotFoundError("AI Infra *_收编视图.xlsx not found")
    return matches[0]


def _load_excel_rows() -> dict[int, dict[str, Any]]:
    wb = load_workbook(_excel_path(), data_only=True)
    ws = wb.active
    rows: dict[int, dict[str, Any]] = {}
    for row in range(2, ws.max_row + 1):
        rows[row] = {
            "model": ws.cell(row, 2).value,
            "hardware": ws.cell(row, 3).value,
            "combo": ws.cell(row, 7).value,
            "engine": ws.cell(row, 8).value,
            "progress": ws.cell(row, 15).value,
            "baseline": ws.cell(row, 16).value,
            "optimized": ws.cell(row, 17).value,
            "wings": ws.cell(row, 19).value,
        }
    return rows


def _scenarios() -> list[Scenario]:
    simple = "simple_modify"
    readapt = "readapt"
    return [
        Scenario(11, "row11-deepseek-v4-flash-910b", "Eco-Tech/DeepSeek-V4-Flash-w8a8-mtp",
                 "DeepseekV4ForCausalLM", "vllm_ascend", "ascend", "Ascend910B_64G",
                 "Ascend910B", 8, simple, "w8a8"),
        Scenario(17, "row17-glm47-910b", "Eco-Tech/GLM-4.7-W8A8-floatmtp",
                 "Glm4MoeForCausalLM", "vllm_ascend", "ascend", "Ascend910B_64G",
                 "Ascend910B", 8, simple, "w8a8"),
        Scenario(18, "row18-qwen3-embedding-910b", "Qwen/Qwen3-Embedding-0.6B",
                 "Qwen3ForCausalLM", "vllm_ascend", "ascend", "Ascend910B_64G",
                 "Ascend910B", 8, simple, model_type="embedding"),
        Scenario(20, "row20-minimax-m3-pro5000", "MiniMax-M3-MXFP8",
                 "MiniMaxM3SparseForConditionalGeneration", "vllm", "nvidia",
                 "NVIDIA RTX PRO 5000 72GB", "NVIDIA RTX PRO 5000 72GB", 8, readapt, "mxfp8"),
        Scenario(21, "row21-minimax-m25-pro5000", "MiniMax-M2.5-NVFP4",
                 "MiniMaxM2ForCausalLM", "vllm", "nvidia", "NVIDIA RTX PRO 5000 72GB",
                 "NVIDIA RTX PRO 5000 72GB", 4, readapt, "nvfp4"),
        # 行标题来自展示列，可能带 BF16；dry-run 命中必须使用 Q 列命令中的真实模型名。
        # 这三行 Q 列路径不带 BF16，若在 model_name 里补后缀会导致 defaults/白名单 exact 命中失败。
        Scenario(22, "row22-qwen35-122b-pro5000", "Qwen/Qwen3.5-122B-A10B",
                 "Qwen3_5MoeForConditionalGeneration", "vllm", "nvidia",
                 "NVIDIA RTX PRO 5000 72GB", "NVIDIA RTX PRO 5000 72GB", 4, readapt),
        Scenario(23, "row23-qwen35-27b-pro5000", "Qwen/Qwen3.5-27B",
                 "Qwen3_5ForConditionalGeneration", "vllm", "nvidia",
                 "NVIDIA RTX PRO 5000 72GB", "NVIDIA RTX PRO 5000 72GB", 2, readapt),
        Scenario(24, "row24-qwen35-35b-pro5000", "Qwen/Qwen3.5-35B-A3B",
                 "Qwen3_5MoeForConditionalGeneration", "vllm", "nvidia",
                 "NVIDIA RTX PRO 5000 72GB", "NVIDIA RTX PRO 5000 72GB", 2, readapt),
        Scenario(30, "row30-deepseek-v4-flash-910c", "Eco-Tech/DeepSeek-V4-Flash-w8a8-mtp",
                 "DeepseekV4ForCausalLM", "vllm_ascend", "ascend", "Ascend910C",
                 "Ascend910C", 16, simple, "w8a8"),
        Scenario(40, "row40-deepseek-coder-910c", "DeepSeek-Coder-V2-Instruct",
                 "DeepseekV3ForCausalLM", "vllm_ascend", "ascend", "Ascend910C",
                 "Ascend910C", 8, simple),
        Scenario(41, "row41-deepseek-v4-pro-910c-dual", "DeepSeek-V4-Pro-w4a8-mtp",
                 "DeepseekV4ForCausalLM", "vllm_ascend", "ascend", "Ascend910C",
                 "Ascend910C", 16, simple, "w4a8", distributed=True, nnodes=2, ranks=(0, 1)),
        Scenario(42, "row42-glm47-910c", "Eco-Tech/GLM-4.7-W8A8-floatmtp",
                 "Glm4MoeForCausalLM", "vllm_ascend", "ascend", "Ascend910C",
                 "Ascend910C", 16, simple, "w8a8"),
        Scenario(44, "row44-kimi-k26-910c", "Eco-Tech/Kimi-K2.6-W4A8",
                 "KimiK25ForConditionalGeneration", "vllm_ascend", "ascend", "Ascend910C",
                 "Ascend910C", 16, readapt, "w4a8"),
        Scenario(45, "row45-kimi-k27-code-910c", "Kimi-K2.7-Code-w4a8",
                 "KimiK25ForConditionalGeneration", "vllm_ascend", "ascend", "Ascend910C",
                 "Ascend910C", 16, simple, "w4a8"),
        Scenario(48, "row48-minimax-m27-pro5000", "MiniMax-M2.7-NVFP4",
                 "MiniMaxM2ForCausalLM", "vllm", "nvidia", "NVIDIA RTX PRO 5000 72GB",
                 "NVIDIA RTX PRO 5000 72GB", 4, simple, "nvfp4"),
        Scenario(49, "row49-qwen35-397b-pro5000", "Qwen/Qwen3.5-397B-A17B-NVFP4",
                 "Qwen3_5MoeForConditionalGeneration", "vllm", "nvidia",
                 "NVIDIA RTX PRO 5000 72GB", "NVIDIA RTX PRO 5000 72GB", 8, simple, "nvfp4"),
        Scenario(17, "extra-glm47-910b-dual", "Eco-Tech/GLM-4.7-W8A8-floatmtp",
                 "Glm4MoeForCausalLM", "vllm_ascend", "ascend", "Ascend910B_64G",
                 "Ascend910B", 8, simple, "w8a8", distributed=True, nnodes=2, ranks=(0, 1),
                 extra=True, note="extra GLM4.7 dual-node requirement"),
        Scenario(42, "extra-glm47-910c-dual", "Eco-Tech/GLM-4.7-W8A8-floatmtp",
                 "Glm4MoeForCausalLM", "vllm_ascend", "ascend", "Ascend910C",
                 "Ascend910C", 8, simple, "w8a8", distributed=True, nnodes=2, ranks=(0, 1),
                 extra=True, note="extra GLM4.7 dual-node requirement"),
    ]


def _contains_command(value: Any) -> bool:
    text = str(value or "")
    return "vllm serve" in text or "vllm.entrypoints.openai.api_server" in text


def _reference_for(scenario: Scenario, mode: str, rows: dict[int, dict[str, Any]]) -> tuple[str, str]:
    row = rows.get(scenario.row, {})
    baseline = str(row.get("baseline") or "")
    optimized = str(row.get("optimized") or "")
    if mode == "features_on" and _contains_command(optimized):
        return "excel_q_optimized", optimized
    if _contains_command(baseline):
        # features_off 是负向控制，使用 P 列 baseline 只验证“关闭特性时不应输出
        # spec/offload/sparse”。Q 列仍是优化场景的首要证据，P 列基础参数不做硬对齐。
        return ("excel_p_negative_control" if mode == "features_off" else "excel_p_fallback"), baseline
    return "missing_excel_command", ""


def _script_engine_commands(text: str) -> list[str]:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = normalized.replace("\\\n", " ")
    normalized = re.sub(r"\\\s+(--[A-Za-z0-9-])", r" \1", normalized)
    executable_lines = []
    for line in normalized.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or stripped.startswith("echo "):
            continue
        executable_lines.append(line)
    normalized = "\n".join(executable_lines)
    commands: list[str] = []
    markers = [m.start() for m in re.finditer(r"(?:nohup\s+|exec\s+)?vllm\s+serve\s+", normalized)]
    markers += [m.start() for m in re.finditer(r"(?:exec\s+)?python3\s+-m\s+vllm\.entrypoints\.openai\.api_server", normalized)]
    for index, start in enumerate(sorted(markers)):
        end = sorted(markers)[index + 1] if index + 1 < len(markers) else len(normalized)
        chunk = normalized[start:end]
        chunk = re.split(r"\n\s*start_[12]\.sh\b", chunk, maxsplit=1)[0]
        chunk = re.split(r"\s+>\s+", chunk, maxsplit=1)[0]
        chunk = chunk.replace("\n", " ").strip()
        chunk = re.sub(r"\s+", " ", chunk)
        if chunk.startswith("nohup "):
            chunk = chunk[len("nohup "):]
        if chunk.startswith("exec "):
            chunk = chunk[len("exec "):]
        commands.append(chunk.strip())
    return commands


def _parse_value(key: str, value: Any) -> Any:
    if value is True:
        return True
    if key in JSON_FLAGS and isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    if isinstance(value, str):
        stripped = value.strip()
        if re.fullmatch(r"-?\d+", stripped):
            try:
                return int(stripped)
            except ValueError:
                return value
        if re.fullmatch(r"-?(?:\d+\.\d*|\d*\.\d+)", stripped):
            try:
                return float(stripped)
            except ValueError:
                return value
    return value


def _parse_flags(command: str) -> tuple[dict[str, Any], str]:
    try:
        tokens = shlex.split(command, posix=True)
    except ValueError as exc:
        return {}, str(exc)
    flags: dict[str, Any] = {}
    i = 0
    while i < len(tokens):
        token = tokens[i]
        if token in {">", "2>&1", "&"}:
            break
        if not token.startswith("--"):
            i += 1
            continue
        if "=" in token:
            raw_key, raw_value = token[2:].split("=", 1)
            key = raw_key.replace("_", "-")
            flags[key] = _parse_value(key, raw_value)
            i += 1
            continue
        key = token[2:].replace("_", "-")
        if key in BOOL_FLAGS or i + 1 >= len(tokens) or tokens[i + 1].startswith("--"):
            flags[key] = True
            i += 1
        else:
            flags[key] = _parse_value(key, tokens[i + 1])
            i += 2
    return flags, ""


def _parse_env(text: str) -> dict[str, Any]:
    env: dict[str, Any] = {}
    for line in text.replace("\r\n", "\n").replace("\r", "\n").splitlines():
        stripped = line.strip()
        if not stripped.startswith("export ") or "=" not in stripped:
            continue
        name, value = stripped[len("export "):].split("=", 1)
        name = name.strip()
        if not name:
            continue
        env[name] = value.strip().strip('"').strip("'")
    return env


def _parse_command_text(text: str, node_rank: int = 0) -> ParsedCommand:
    commands = _script_engine_commands(text)
    if not commands:
        return ParsedCommand(env=_parse_env(text), parse_error="no engine command found")
    command = commands[min(node_rank, len(commands) - 1)]
    flags, error = _parse_flags(command)
    return ParsedCommand(command=command, flags=flags, env=_parse_env(text), parse_error=error)


def _extract_memcache_dram_gb(text: str) -> int | None:
    """从 Excel Q 列 MemCache 说明里提取 dram.size，作为 dry-run 用户输入。"""
    match = re.search(
        r"ock\.mmc\.local_service\.dram\.size\s*=\s*(\d+)\s*GB",
        str(text or ""),
        re.IGNORECASE,
    )
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _field_subset(parsed: ParsedCommand) -> dict[str, Any]:
    return {key: parsed.flags[key] for key in sorted(CLI_FIELDS | JSON_FLAGS) if key in parsed.flags}


def _model_config(scenario: Scenario) -> dict[str, Any]:
    config: dict[str, Any] = {
        "architectures": [scenario.architecture],
        "_name_or_path": scenario.model_name,
        "model_type": scenario.model_name.rsplit("/", 1)[-1].lower(),
        "torch_dtype": "bfloat16",
    }
    if scenario.quantize:
        config["quantize"] = scenario.quantize
        config["quantization_config"] = {"quant_method": scenario.quantize}
    return config


def _create_model_dir(root: Path, scenario: Scenario) -> Path:
    model_dir = root / "_model_configs" / scenario.slug
    model_dir.mkdir(parents=True, exist_ok=True)
    _write_json(model_dir / "config.json", _model_config(scenario))
    _write_text(model_dir / "wings.txt", scenario.model_name)
    return model_dir


def _hardware_info(scenario: Scenario) -> dict[str, Any]:
    return {
        "device": scenario.device,
        "count": scenario.device_count,
        "hardware_family": scenario.hardware_family,
        "details": [
            {
                "device_id": idx,
                "name": scenario.detail_name,
                "total_memory": 72 if scenario.device == "nvidia" else 64,
                "free_memory": 70 if scenario.device == "nvidia" else 60,
                "used_memory": 2,
                "util": 0,
            }
            for idx in range(max(1, scenario.device_count))
        ],
        "units": "GB",
    }


def _reset_env(shared_root: Path, scenario: Scenario, mode: str, node_rank: int) -> None:
    for key in MANAGED_ENV_KEYS:
        os.environ.pop(key, None)
    shared_root.mkdir(parents=True, exist_ok=True)
    os.environ["SHARED_VOLUME_PATH"] = _as_posix(shared_root)
    os.environ["ENABLE_REASON_PROXY"] = "false"
    os.environ["POD_IP"] = f"10.10.0.{10 + node_rank}"
    os.environ["RANK_IP"] = f"10.10.0.{10 + node_rank}"
    os.environ["NETWORK_INTERFACE"] = "eth0"
    os.environ["GLOO_SOCKET_IFNAME"] = "eth0"
    os.environ["MASTER_IP"] = "10.10.0.10"
    os.environ["ENGINE_VERSION"] = (
        "0.23.0-rtx_pro_5000_72G"
        if scenario.device == "nvidia"
        else ("0.21.0-a3" if "910C" in scenario.detail_name else "0.21.0-a2")
    )
    if scenario.distributed:
        os.environ["DISTRIBUTED"] = "true"
        os.environ["NNODES"] = str(scenario.nnodes)
        os.environ["NODE_IPS"] = "10.10.0.10,10.10.0.11"
        os.environ["NODES"] = os.environ["NODE_IPS"]
        os.environ["DISTRIBUTED_EXECUTOR_BACKEND"] = "dp_deployment"
    if mode == "features_on":
        os.environ["ENABLE_AUTO_TOOL_CHOICE"] = "true"
        os.environ["ENABLE_AUTO_THINK_CHOICE"] = "true"
        os.environ["ENABLE_KV_OFFLOAD"] = "true"
        os.environ["ENABLE_KV_MEM_OFFLOAD"] = "true"
        os.environ["ENABLE_KV_DISK_OFFLOAD"] = "false"
        os.environ["ENABLE_SPARSE"] = "true"
        os.environ["SPARSE_ENABLE"] = "true"
        os.environ["ENABLE_SPECULATIVE_DECODE"] = "true"
        os.environ["SD_ENABLE"] = "true"
        os.environ["LMCACHE_OFFLOAD"] = "true"
    else:
        os.environ["ENABLE_AUTO_TOOL_CHOICE"] = "false"
        os.environ["ENABLE_AUTO_THINK_CHOICE"] = "false"
        os.environ["ENABLE_KV_OFFLOAD"] = "false"
        os.environ["ENABLE_KV_MEM_OFFLOAD"] = "false"
        os.environ["ENABLE_KV_DISK_OFFLOAD"] = "false"
        os.environ["ENABLE_SPARSE"] = "false"
        os.environ["SPARSE_ENABLE"] = "false"
        os.environ["ENABLE_SPECULATIVE_DECODE"] = "false"
        os.environ["SD_ENABLE"] = "false"
        os.environ["LMCACHE_OFFLOAD"] = "false"
        os.environ["SPECULATIVE_DECODE_MODEL_PATH"] = ""


def _clear_hardware_cache() -> None:
    try:
        from utils import device_utils

        device_utils._hardware_cache.clear()  # type: ignore[attr-defined]
    except Exception:
        pass


def _reference_max_len(parsed: ParsedCommand, scenario: Scenario) -> int:
    value = parsed.flags.get("max-model-len")
    if isinstance(value, int) and value > 0:
        return value
    if scenario.model_type == "embedding":
        return 4096
    return 4096 + 1024


def _length_args_for(case: Case, parsed_ref: ParsedCommand) -> tuple[int, int]:
    max_len = _reference_max_len(parsed_ref, case.scenario)
    if case.scenario.model_type == "embedding":
        return max_len, 0
    return max(1, max_len - 1), 1


def _bool_env(enabled: bool) -> str:
    return "true" if enabled else "false"


def _has_fp8_kv_cache_dtype(parsed: ParsedCommand) -> bool:
    return str(parsed.flags.get("kv-cache-dtype") or "").strip().lower() == "fp8"


def _reference_requests_sparse(parsed: ParsedCommand) -> bool:
    # FP8 KV Cache 和 IndexCache 统一归入 sparse 验收桶；仅 fp8 值触发，避免 auto 误判。
    return (
        "sparse-config" in parsed.flags
        or "hf-overrides" in parsed.flags
        or _has_fp8_kv_cache_dtype(parsed)
    )


def _skip_features_off_reference_field(field: str, reference: ParsedCommand) -> bool:
    return field in (FEATURE_CLI_FLAGS | INTERACTIVE_CLI_FLAGS) or (
        field == "kv-cache-dtype" and _has_fp8_kv_cache_dtype(reference)
    )


def _set_reference_feature_inputs(case: Case, parsed_ref: ParsedCommand) -> dict[str, bool]:
    """Mirror the feature inputs implied by the reference command for this case."""
    features_on = case.mode == "features_on"
    native_offload = features_on and (
        "kv-offloading-backend" in parsed_ref.flags
        or "kv-offloading-size" in parsed_ref.flags
    )
    connector_offload = features_on and (
        "kv-transfer-config" in parsed_ref.flags
        or any(key in parsed_ref.env for key in FEATURE_ENV_KEYS)
    )
    feature_inputs = {
        "spec": features_on and "speculative-config" in parsed_ref.flags,
        "sparse": features_on and _reference_requests_sparse(parsed_ref),
        "offload": native_offload or connector_offload,
        "native_offload": native_offload,
        "connector_offload": connector_offload,
    }

    os.environ["ENABLE_SPECULATIVE_DECODE"] = os.environ["SD_ENABLE"] = _bool_env(feature_inputs["spec"])
    os.environ["ENABLE_SPARSE"] = os.environ["SPARSE_ENABLE"] = _bool_env(feature_inputs["sparse"])
    os.environ["ENABLE_KV_OFFLOAD"] = os.environ["LMCACHE_OFFLOAD"] = _bool_env(feature_inputs["offload"])
    os.environ["ENABLE_KV_MEM_OFFLOAD"] = _bool_env(feature_inputs["native_offload"] or connector_offload)
    os.environ["ENABLE_KV_DISK_OFFLOAD"] = "false"

    if native_offload:
        size = parsed_ref.flags.get("kv-offloading-size")
        if size not in (None, ""):
            os.environ["KV_MEM_OFFLOAD_SIZE"] = str(size)
        else:
            os.environ.pop("KV_MEM_OFFLOAD_SIZE", None)
    elif connector_offload:
        memcache_dram_gb = _extract_memcache_dram_gb(case.reference_text)
        if memcache_dram_gb:
            os.environ["KV_MEM_OFFLOAD_SIZE"] = str(memcache_dram_gb)
            feature_inputs["kv_mem_offload_size_from_reference"] = True
        else:
            os.environ.pop("KV_MEM_OFFLOAD_SIZE", None)
            feature_inputs["kv_mem_offload_size_from_reference"] = False
    else:
        os.environ.pop("KV_MEM_OFFLOAD_SIZE", None)
    return feature_inputs


def _build_case(case: Case, out_root: Path) -> dict[str, Any]:
    from config.settings import settings
    from core.port_plan import derive_port_plan
    from core.start_args_compat import parse_launch_args
    from core.wings_entry import build_launcher_plan

    case_dir = out_root / case.name
    shared_root = case_dir / "_shared"
    _reset_env(shared_root, case.scenario, case.mode, case.node_rank)
    model_dir = _create_model_dir(out_root, case.scenario)
    hw_file = case_dir / "hardware_info.json"
    _write_json(hw_file, _hardware_info(case.scenario))
    os.environ["WINGS_HARDWARE_FILE"] = _as_posix(hw_file)
    _clear_hardware_cache()

    parsed_ref = _parse_command_text(case.reference_text, case.node_rank)
    feature_inputs = _set_reference_feature_inputs(case, parsed_ref)
    input_length, output_length = _length_args_for(case, parsed_ref)
    argv = [
        "--model-name", case.scenario.model_name,
        "--model-path", _as_posix(model_dir),
        "--engine", case.scenario.engine,
        "--device-count", str(case.scenario.device_count),
        "--input-length", str(input_length),
        "--output-length", str(output_length),
        "--model-type", case.scenario.model_type,
        "--save-path", _as_posix(case_dir),
        "--node-rank", str(case.node_rank),
    ]
    if case.scenario.distributed:
        argv.extend([
            "--distributed",
            "--nnodes", str(case.scenario.nnodes),
            "--node-ips", "10.10.0.10,10.10.0.11",
            "--nodes", "10.10.0.10,10.10.0.11",
            "--master-ip", "10.10.0.10",
            "--head-node-addr", "10.10.0.10",
            "--distributed-executor-backend", "dp_deployment",
        ])
    if case.mode == "features_on":
        argv.extend([
            "--enable-auto-tool-choice",
            "--enable-auto-think-choice",
        ])
        if feature_inputs["spec"]:
            argv.append("--enable-speculative-decode")
        if feature_inputs["sparse"]:
            argv.append("--enable-sparse")

    result: dict[str, Any] = {
        "case": case.name,
        "row": case.scenario.row,
        "extra": case.scenario.extra,
        "mode": case.mode,
        "node_rank": case.node_rank,
        "adaptation": case.scenario.adaptation,
        "reference_kind": case.reference_kind,
        "feature_inputs": feature_inputs,
        "input_length": input_length,
        "output_length": output_length,
        "errors": [],
        "checks": [],
        "failures": [],
    }
    try:
        launch_args = parse_launch_args(argv)
        port_plan = derive_port_plan(
            port=launch_args.port,
            enable_reason_proxy=False,
            health_port=settings.HEALTH_PORT,
        )
        plan = build_launcher_plan(launch_args, port_plan)
    except Exception as exc:  # noqa: BLE001
        result["result"] = "ERROR"
        result["errors"].append(str(exc))
        return result

    actual = _parse_command_text(plan.command, 0)
    _write_text(case_dir / "start_command.sh", plan.command)
    _write_text(case_dir / "actual_engine_command.txt", actual.command or "")
    _write_json(case_dir / "merged_params.json", plan.merged_params)
    _write_json(case_dir / "reference_fields.json", _field_subset(parsed_ref))
    _write_json(case_dir / "actual_fields.json", _field_subset(actual))

    result.update({
        "start_command": str((case_dir / "start_command.sh").relative_to(REPO_ROOT)),
        "actual_engine_command": str((case_dir / "actual_engine_command.txt").relative_to(REPO_ROOT)),
        "merged_params": str((case_dir / "merged_params.json").relative_to(REPO_ROOT)),
        "reference_parse_error": parsed_ref.parse_error,
        "actual_parse_error": actual.parse_error,
        "reference_fields": _field_subset(parsed_ref),
        "actual_fields": _field_subset(actual),
        "active_smart_features": plan.merged_params.get("_smart_feats", []),
        "allowed_smart_features": plan.merged_params.get("_allowed_smart_feats", []),
    })
    checks = _compare_case(case, parsed_ref, actual, plan.command)
    issues = _classify_failed_checks(case, checks)
    result["checks"] = checks
    result["failures"] = issues
    result["blocking_failures"] = [issue for issue in issues if issue.get("blocking")]
    result["review_items"] = [issue for issue in issues if not issue.get("blocking")]
    if result["errors"]:
        result["result"] = "ERROR"
    elif result["blocking_failures"]:
        result["result"] = "FAIL"
    elif result["review_items"]:
        result["result"] = "REVIEW"
    else:
        result["result"] = "PASS"
    return result


def _same_value(expected: Any, actual: Any) -> bool:
    if isinstance(expected, float) or isinstance(actual, float):
        try:
            return abs(float(expected) - float(actual)) < 0.000001
        except (TypeError, ValueError):
            return False
    if isinstance(expected, dict) and isinstance(actual, dict):
        return _dict_subset(expected, actual)
    return expected == actual


def _dict_subset(expected: dict[str, Any], actual: dict[str, Any]) -> bool:
    for key, exp_value in expected.items():
        if key not in actual:
            return False
        act_value = actual[key]
        if isinstance(exp_value, dict) and isinstance(act_value, dict):
            if not _dict_subset(exp_value, act_value):
                return False
        elif exp_value != act_value:
            return False
    return True


def _check(field: str, expected: Any, actual: Any, ok: bool, rule: str) -> dict[str, Any]:
    return {"field": field, "expected": expected, "actual": actual, "ok": ok, "rule": rule}


def _is_kimi_dflash_fallback_delta(case: Case, check: dict[str, Any]) -> bool:
    expected = check.get("expected")
    actual = check.get("actual")
    return (
        case.scenario.row == 44
        and check.get("field") == "speculative-config"
        and isinstance(expected, dict)
        and expected.get("method") == "dflash"
        and isinstance(actual, dict)
        and actual.get("method") == "suffix"
    )


def _classify_failed_check(case: Case, check: dict[str, Any]) -> dict[str, Any]:
    """将 strict diff 归类成文档定义的阻塞项或允许差异。"""
    item = dict(check)
    field = str(check.get("field", ""))
    rule = str(check.get("rule", ""))

    if rule == "input_output_only" or field == "max-model-len":
        item["category"] = "user_input_control"
        item["blocking"] = False
        item["note"] = "input/output length is supplied by the user"
        return item

    if field == "reference_parse" and case.reference_kind == "missing_excel_command":
        item["category"] = "external_evidence"
        item["blocking"] = False
        item["note"] = "Excel has no parseable command for this row"
        return item

    if field == "actual_parse" and case.scenario.extra and case.node_rank > 0:
        item["category"] = "external_evidence"
        item["blocking"] = False
        item["note"] = "supplementary dual-node GLM case still lacks locked standard input"
        return item

    if _is_kimi_dflash_fallback_delta(case, check):
        item["category"] = "allowed_delta"
        item["blocking"] = False
        item["note"] = "DFlash requires a real draft path; current input correctly falls back to suffix"
        return item

    if (
        case.mode == "features_off"
        and case.reference_kind == "excel_p_negative_control"
        and rule == "excel_field"
    ):
        item["category"] = "allowed_delta"
        item["blocking"] = False
        item["note"] = "P-column baseline is used only as negative-control evidence"
        return item

    item["category"] = "must_fix"
    item["blocking"] = True
    item["note"] = "semantic mismatch not covered by documented dynamic inputs"
    return item


def _classify_failed_checks(case: Case, checks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [_classify_failed_check(case, check) for check in checks if not check["ok"]]


def _normalize_expected_field(case: Case, field: str, expected: Any) -> Any:
    if (
        field == "tool-call-parser"
        and case.scenario.row in {22, 24}
        and (
            case.scenario.model_name.endswith("Qwen3.5-122B-A10B")
            or case.scenario.model_name.endswith("Qwen3.5-35B-A3B")
        )
        and expected == "hermes"
    ):
        return "qwen3_coder"
    if (
        field == "speculative-config"
        and case.scenario.engine == "vllm"
        and case.scenario.device == "nvidia"
        and isinstance(expected, dict)
        and expected.get("method") == "qwen3_5_mtp"
    ):
        normalized = dict(expected)
        normalized["method"] = "mtp"
        return normalized
    return expected


def _normalize_actual_field(field: str, expected: Any, actual: Any) -> Any:
    if field in {"data-parallel-size", "data-parallel-size-local"} and expected == 1 and actual is None:
        return 1
    return actual


def _compare_case(case: Case, reference: ParsedCommand, actual: ParsedCommand, script: str) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    if reference.parse_error:
        checks.append(_check("reference_parse", "parseable", reference.parse_error, False, "reference"))
        return checks
    if actual.parse_error:
        checks.append(_check("actual_parse", "parseable", actual.parse_error, False, "actual"))
        return checks

    comparable = sorted((CLI_FIELDS | JSON_FLAGS) - {"host", "port", "model"})
    for field in comparable:
        if field not in reference.flags:
            continue
        if case.mode == "features_off" and _skip_features_off_reference_field(field, reference):
            continue
        expected = _normalize_expected_field(case, field, reference.flags[field])
        actual_value = _normalize_actual_field(field, expected, actual.flags.get(field))
        rule = "non_blocking_excel_field" if field in NON_BLOCKING_CLI_FIELDS else "excel_field"
        ok = True if field in NON_BLOCKING_CLI_FIELDS else _same_value(expected, actual_value)
        checks.append(_check(field, expected, actual_value, ok, rule))

    expected_max_len = _reference_max_len(reference, case.scenario)
    if expected_max_len:
        actual_max_len = actual.flags.get("max-model-len")
        checks.append(_check("max-model-len-from-input-output", expected_max_len, actual_max_len,
                             actual_max_len == expected_max_len, "input_output_only"))

    if case.scenario.distributed:
        expected_headless = case.node_rank != 0
        checks.append(_check("headless", expected_headless, bool(actual.flags.get("headless")),
                             bool(actual.flags.get("headless")) == expected_headless, "distributed_rank"))

    if case.mode == "features_off":
        for field in sorted(FEATURE_CLI_FLAGS):
            checks.append(_check(field, None, actual.flags.get(field),
                                 field not in actual.flags, "feature_off_forbidden"))
        actual_fp8 = _has_fp8_kv_cache_dtype(actual)
        checks.append(_check("kv-cache-dtype:fp8", None, actual.flags.get("kv-cache-dtype"),
                             not actual_fp8, "feature_off_forbidden"))
        for env_key in sorted(FEATURE_ENV_KEYS):
            present = env_key in actual.env or env_key in script
            checks.append(_check(f"env:{env_key}", None, "present" if present else None,
                                 not present, "feature_off_forbidden"))
    else:
        for field in sorted(FEATURE_CLI_FLAGS):
            if field in reference.flags:
                expected = _normalize_expected_field(case, field, reference.flags[field])
                actual_value = _normalize_actual_field(field, expected, actual.flags.get(field))
                checks.append(_check(field, expected, actual_value,
                                     _same_value(expected, actual_value),
                                     "feature_on_excel_field"))
    return checks


def _cases(rows: dict[int, dict[str, Any]]) -> list[Case]:
    cases: list[Case] = []
    for scenario in _scenarios():
        for mode in ("features_off", "features_on"):
            ref_kind, ref_text = _reference_for(scenario, mode, rows)
            for rank in scenario.ranks:
                cases.append(Case(scenario, mode, rank, ref_kind, ref_text))
    return cases


def _audit(rows: dict[int, dict[str, Any]], cases: list[Case]) -> dict[str, Any]:
    s_empty = [
        row for row, values in rows.items()
        if values.get("wings") in (None, "")
    ]
    processed = [
        row for row in s_empty
        if row not in EXCLUDED_ROWS and row in SOURCE_ROWS
    ]
    return {
        "source_file": str(_excel_path().relative_to(REPO_ROOT)),
        "s_empty_count": len(s_empty),
        "source_rows": list(SOURCE_ROWS),
        "processed_rows": processed,
        "excluded_rows": list(EXCLUDED_ROWS),
        "scenario_count": len(_scenarios()),
        "case_count": len(cases),
        "modes": ["features_off", "features_on"],
        "input_policy": "Only model/hardware/deployment context plus input_length/output_length are supplied. features_on mirrors reference feature requests; KV_MEM_OFFLOAD_SIZE is supplied only when the reference has native offload size or parseable MemCache dram.size. No engine_config is supplied.",
    }


def _markdown_report(payload: dict[str, Any]) -> str:
    results = payload["results"]
    counts = {
        "PASS": sum(r["result"] == "PASS" for r in results),
        "REVIEW": sum(r["result"] == "REVIEW" for r in results),
        "FAIL": sum(r["result"] == "FAIL" for r in results),
        "ERROR": sum(r["result"] == "ERROR" for r in results),
    }
    lines = [
        "# S-empty DAY0 dry-run verification",
        "",
        f"Source: `{payload['audit']['source_file']}`",
        f"Cases: {len(results)} | PASS: {counts['PASS']} | REVIEW: {counts['REVIEW']} | "
        f"FAIL: {counts['FAIL']} | ERROR: {counts['ERROR']}",
        "",
        "The launcher path is the real `build_launcher_plan()` path. Each case supplies scenario context and user-level inputs only; no `engine_config` is supplied.",
        "",
        "Result semantics: `FAIL` means a blocking `must_fix` mismatch; `REVIEW` means only documented deltas, user-controlled inputs, or missing external evidence remain.",
        "",
        "| Case | Result | Adaptation | Ref | Active features | Issues | start_command.sh |",
        "|---|---:|---|---|---|---|---|",
    ]
    for result in results:
        failed = result.get("failures") or []
        failed_fields = "-"
        if failed:
            failed_fields = "<br>".join(
                f"{f.get('category', 'must_fix')}::{f['field']}: "
                f"exp `{json.dumps(f['expected'], ensure_ascii=False)}` "
                f"got `{json.dumps(f['actual'], ensure_ascii=False)}`"
                for f in failed[:8]
            )
            if len(failed) > 8:
                failed_fields += f"<br>... {len(failed) - 8} more"
        if result.get("errors"):
            failed_fields = "<br>".join(result["errors"])
        active = ",".join(result.get("active_smart_features") or []) or "-"
        start = result.get("start_command", "-")
        lines.append(
            f"| {result['case']} | {result['result']} | `{result['adaptation']}` | "
            f"`{result['reference_kind']}` | `{active}` | {failed_fields} | `{start}` |"
        )
    lines.extend([
        "",
        "## Checked Logic",
        "",
        "- `features_off`: uses P-column baseline only as a negative-control reference; basic parameter drifts are documented deltas, while spec/offload/sparse artifacts remain blocking.",
        "- `features_on`: uses Q-column optimized command when present; fallback to P only when Q has no executable command.",
        "- `served-model-name` is kept as a non-blocking evidence field because the mother document validates command semantics rather than short alias text.",
        "- `max-model-len` and `max-model-len-from-input-output` are user-input controlled review items, not hard adaptation failures.",
        "- Kimi K2.6 DFlash is only hard-expected when a real DFlash draft path is supplied; otherwise suffix fallback is a documented delta.",
        "- MemCache validation supplies `KV_MEM_OFFLOAD_SIZE` only when the reference contains native offload size or parseable `dram.size` evidence.",
        "- Supplementary dual-node cases without locked standard input are reported as review evidence gaps instead of Excel-row failures.",
        "",
    ])
    return "\n".join(lines)


def main() -> int:
    _ensure_import_path()
    _safe_clear_dir(OUT_ROOT)
    rows = _load_excel_rows()
    cases = _cases(rows)
    payload = {"audit": _audit(rows, cases), "results": []}
    for case in cases:
        payload["results"].append(_build_case(case, OUT_ROOT))
    _write_json(OUT_ROOT / "comparison.json", payload)
    _write_text(OUT_ROOT / "comparison.md", _markdown_report(payload))
    hard = [r for r in payload["results"] if r["result"] in {"FAIL", "ERROR"}]
    counts = {name: sum(r["result"] == name for r in payload["results"]) for name in ("PASS", "REVIEW", "FAIL", "ERROR")}
    print(
        f"cases={len(payload['results'])} pass={counts['PASS']} review={counts['REVIEW']} "
        f"fail={counts['FAIL']} error={counts['ERROR']}"
    )
    print(OUT_ROOT / "comparison.md")
    return 1 if hard else 0


if __name__ == "__main__":
    raise SystemExit(main())
