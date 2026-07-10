#!/usr/bin/env python3
"""生成并校验 Qwen Day0 dry-run 启动命令。

该校验器以 ``qwen模型收编.xlsx`` 作为标准工作簿，通过
``build_launcher_plan`` 走当前代码路径生成真实的 ``start_command.sh``。
这里故意校验最终 vLLM 命令，而不是只读取默认 JSON，因为白名单门控、特性开关、
硬件识别、MemCache 前置命令渲染和 fallback 命令生成都发生在配置选择之后。
"""

from __future__ import annotations

import json
import os
import shlex
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
WINGS_CONTROL_ROOT = REPO_ROOT / "wings_control"
DOC_ROOT = WINGS_CONTROL_ROOT / "docs" / "DAY0"
DRY_RUN_ROOT = DOC_ROOT / "dry_run"
STANDARD_OUT_ROOT = DRY_RUN_ROOT / "qwen_day0_current"
REUSE_OUT_ROOT = DRY_RUN_ROOT / "qwen_day0_910b_reuse_current"
EXCEL_STANDARD = DOC_ROOT / "qwen\u6a21\u578b\u6536\u7f16.xlsx"
EXCEL_BASELINE = DOC_ROOT / "qwen\u4f18\u5316\u53c2\u6570\u9002\u914d\u57fa\u51c6.xlsx"


QWEN35_ADDITIONAL = {
    "enable_cpu_binding": False,
    "ascend_compilation_config": {
        "enable_npugraph_ex": True,
        "enable_static_kernel": False,
    },
    "multistream_overlap_shared_expert": False,
}
QWEN36_ADDITIONAL = {
    "enable_cpu_binding": True,
    "ascend_compilation_config": {
        "enable_npugraph_ex": True,
    },
}
COMPILATION_CONFIG = {"cudagraph_mode": "FULL_DECODE_ONLY"}
QWEN_MEMCACHE_KV_CONFIG = {
    "kv_connector": "AscendStoreConnector",
    "kv_role": "kv_both",
    "kv_connector_extra_config": {
        "lookup_rpc_port": "0",
        "backend": "memcache",
    },
}


@dataclass(frozen=True)
class Scenario:
    name: str
    model_name: str
    architecture: str
    hardware_family: str
    device_count: int
    port: int
    tp: int
    dp: int
    max_num_seqs: int
    max_model_len: int
    max_num_batched_tokens: int
    gpu_memory_utilization: float
    served_model_name: str
    mtp_tokens: int
    additional_config: dict[str, Any]
    source_row: int
    source_column: str
    source_cell_refs: tuple[str, ...]
    quantization: str | None = None
    enable_expert_parallel: bool = False
    offload: bool = False
    memcache_ports: tuple[int, int] | None = None
    language_model_only: bool = False
    seed: int = 1024
    is_reuse: bool = False
    notes: str = ""


def _dense(
    *,
    name: str,
    model_name: str,
    card: str,
    device_count: int,
    port: int,
    tp: int,
    max_num_seqs: int,
    max_model_len: int,
    max_num_batched_tokens: int,
    gpu_memory_utilization: float,
    served_model_name: str,
    mtp_tokens: int,
    source_row: int,
    source_column: str,
    source_cell_refs: tuple[str, ...],
    qwen36: bool = False,
    quantization: str | None = None,
    offload: bool = False,
    memcache_ports: tuple[int, int] | None = None,
    is_reuse: bool = False,
    notes: str = "",
) -> Scenario:
    return Scenario(
        name=name,
        model_name=model_name,
        architecture="Qwen3_5ForConditionalGeneration",
        hardware_family="Ascend910B_64G" if card == "910b" else "Ascend910C",
        device_count=device_count,
        port=port,
        tp=tp,
        dp=1,
        max_num_seqs=max_num_seqs,
        max_model_len=max_model_len,
        max_num_batched_tokens=max_num_batched_tokens,
        gpu_memory_utilization=gpu_memory_utilization,
        served_model_name=served_model_name,
        mtp_tokens=mtp_tokens,
        additional_config=QWEN36_ADDITIONAL if qwen36 else QWEN35_ADDITIONAL,
        quantization=quantization,
        offload=offload,
        memcache_ports=memcache_ports,
        language_model_only=not qwen36,
        source_row=source_row,
        source_column=source_column,
        source_cell_refs=source_cell_refs,
        is_reuse=is_reuse,
        notes=notes,
    )


def _moe(
    *,
    name: str,
    model_name: str,
    card: str,
    device_count: int,
    port: int,
    tp: int,
    max_num_seqs: int,
    max_model_len: int,
    max_num_batched_tokens: int,
    gpu_memory_utilization: float,
    served_model_name: str,
    mtp_tokens: int,
    source_row: int,
    source_column: str,
    source_cell_refs: tuple[str, ...],
    qwen36: bool = False,
    quantization: str | None = None,
    offload: bool = False,
    memcache_ports: tuple[int, int] | None = None,
    is_reuse: bool = False,
    notes: str = "",
) -> Scenario:
    return Scenario(
        name=name,
        model_name=model_name,
        architecture="Qwen3_5MoeForConditionalGeneration",
        hardware_family="Ascend910B_64G" if card == "910b" else "Ascend910C",
        device_count=device_count,
        port=port,
        tp=tp,
        dp=1,
        max_num_seqs=max_num_seqs,
        max_model_len=max_model_len,
        max_num_batched_tokens=max_num_batched_tokens,
        gpu_memory_utilization=gpu_memory_utilization,
        served_model_name=served_model_name,
        mtp_tokens=mtp_tokens,
        additional_config=QWEN36_ADDITIONAL if qwen36 else QWEN35_ADDITIONAL,
        quantization=quantization,
        enable_expert_parallel=True,
        offload=offload,
        memcache_ports=memcache_ports,
        language_model_only=not qwen36,
        source_row=source_row,
        source_column=source_column,
        source_cell_refs=source_cell_refs,
        is_reuse=is_reuse,
        notes=notes,
    )


STANDARD_SCENARIOS: list[Scenario] = [
    _dense(
        name="Qwen3.5-27B-910C",
        model_name="Qwen/Qwen3.5-27B",
        card="910c",
        device_count=2,
        port=7878,
        tp=2,
        max_num_seqs=8,
        max_model_len=131072,
        max_num_batched_tokens=16384,
        gpu_memory_utilization=0.75,
        served_model_name="qwen35",
        mtp_tokens=1,
        offload=True,
        memcache_ports=(50051, 50061),
        source_row=2,
        source_column="910C optimized",
        source_cell_refs=("G2", "J2"),
    ),
    _dense(
        name="Qwen3.5-27B-910B",
        model_name="Qwen/Qwen3.5-27B",
        card="910b",
        device_count=4,
        port=7898,
        tp=4,
        max_num_seqs=8,
        max_model_len=8192,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.75,
        served_model_name="qwen35",
        mtp_tokens=3,
        source_row=2,
        source_column="910B optimized",
        source_cell_refs=("I2",),
    ),
    _moe(
        name="Qwen3.5-35B-A3B-910C",
        model_name="Qwen/Qwen3.5-35B-A3B",
        card="910c",
        device_count=2,
        port=7897,
        tp=2,
        max_num_seqs=32,
        max_model_len=161072,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen35",
        mtp_tokens=1,
        source_row=3,
        source_column="910C optimized",
        source_cell_refs=("G3",),
    ),
    _moe(
        name="Qwen3.5-122B-A10B-910C",
        model_name="Qwen/Qwen3.5-122B-A10B",
        card="910c",
        device_count=8,
        port=6901,
        tp=8,
        max_num_seqs=8,
        max_model_len=131072,
        max_num_batched_tokens=16384,
        gpu_memory_utilization=0.9,
        served_model_name="qwen35",
        mtp_tokens=1,
        source_row=4,
        source_column="910C optimized",
        source_cell_refs=("G4",),
    ),
    _moe(
        name="Qwen3.5-397B-A17B-910C",
        model_name="Qwen/Qwen3.5-397B-A17B",
        card="910c",
        device_count=16,
        port=6901,
        tp=16,
        max_num_seqs=4,
        max_model_len=16384,
        max_num_batched_tokens=16384,
        gpu_memory_utilization=0.92,
        served_model_name="qwen35",
        mtp_tokens=1,
        source_row=5,
        source_column="910C optimized",
        source_cell_refs=("G5",),
    ),
    _dense(
        name="Qwen3.6-27B-910C",
        model_name="Qwen/Qwen3.6-27B",
        card="910c",
        device_count=2,
        port=7799,
        tp=2,
        max_num_seqs=32,
        max_model_len=131072,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen36",
        mtp_tokens=3,
        offload=True,
        memcache_ports=(50071, 50081),
        qwen36=True,
        source_row=6,
        source_column="910C optimized + dependency script",
        source_cell_refs=("G6", "J6"),
    ),
    _dense(
        name="Qwen3.6-27B-w8a8-910C",
        model_name="Eco-Tech/Qwen3.6-27B-w8a8",
        card="910c",
        device_count=2,
        port=7799,
        tp=2,
        max_num_seqs=32,
        max_model_len=131072,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen36",
        mtp_tokens=3,
        quantization="ascend",
        offload=True,
        memcache_ports=(50071, 50081),
        qwen36=True,
        source_row=7,
        source_column="910C optimized + dependency script",
        source_cell_refs=("G7", "J6"),
    ),
    _dense(
        name="Qwen3.6-27B-w8a8-910B",
        model_name="Eco-Tech/Qwen3.6-27B-w8a8",
        card="910b",
        device_count=4,
        port=7899,
        tp=4,
        max_num_seqs=32,
        max_model_len=131702,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen36",
        mtp_tokens=3,
        quantization="ascend",
        qwen36=True,
        source_row=7,
        source_column="910B optimized",
        source_cell_refs=("I7",),
    ),
    _moe(
        name="Qwen3.6-35B-A3B-910C",
        model_name="Qwen/Qwen3.6-35B-A3B",
        card="910c",
        device_count=2,
        port=7799,
        tp=2,
        max_num_seqs=32,
        max_model_len=131072,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen36",
        mtp_tokens=3,
        offload=True,
        memcache_ports=(50071, 50081),
        qwen36=True,
        source_row=8,
        source_column="910C optimized + dependency script",
        source_cell_refs=("G8", "J6"),
    ),
    _moe(
        name="Qwen3.6-35B-A3B-w8a8-910C",
        model_name="Eco-Tech/Qwen3.6-35B-A3B-w8a8",
        card="910c",
        device_count=2,
        port=7799,
        tp=2,
        max_num_seqs=32,
        max_model_len=131072,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen36",
        mtp_tokens=3,
        quantization="ascend",
        offload=True,
        memcache_ports=(50071, 50081),
        qwen36=True,
        source_row=9,
        source_column="910C optimized + dependency script",
        source_cell_refs=("G9", "J6"),
    ),
    _moe(
        name="Qwen3.6-35B-A3B-w8a8-910B",
        model_name="Eco-Tech/Qwen3.6-35B-A3B-w8a8",
        card="910b",
        device_count=4,
        port=7899,
        tp=4,
        max_num_seqs=32,
        max_model_len=131702,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen36",
        mtp_tokens=3,
        quantization="ascend",
        qwen36=True,
        source_row=9,
        source_column="910B optimized",
        source_cell_refs=("I9",),
    ),
]

REUSE_SCENARIOS: list[Scenario] = [
    _moe(
        name="Qwen3.5-35B-A3B-910B-reuse",
        model_name="Qwen/Qwen3.5-35B-A3B",
        card="910b",
        device_count=2,
        port=7897,
        tp=2,
        max_num_seqs=32,
        max_model_len=161072,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen35",
        mtp_tokens=1,
        source_row=3,
        source_column="910B reuse of row3 910C optimized",
        source_cell_refs=("G3",),
        is_reuse=True,
        notes="910B row is not measured in the source workbook; config intentionally reuses 910C.",
    ),
    _moe(
        name="Qwen3.5-122B-A10B-910B-reuse",
        model_name="Qwen/Qwen3.5-122B-A10B",
        card="910b",
        device_count=8,
        port=6901,
        tp=8,
        max_num_seqs=8,
        max_model_len=131072,
        max_num_batched_tokens=16384,
        gpu_memory_utilization=0.9,
        served_model_name="qwen35",
        mtp_tokens=1,
        source_row=4,
        source_column="910B reuse of row4 910C optimized",
        source_cell_refs=("G4",),
        is_reuse=True,
        notes="910B row is not measured in the source workbook; config intentionally reuses 910C.",
    ),
    _dense(
        name="Qwen3.6-27B-910B-reuse",
        model_name="Qwen/Qwen3.6-27B",
        card="910b",
        device_count=2,
        port=7799,
        tp=2,
        max_num_seqs=32,
        max_model_len=131072,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen36",
        mtp_tokens=3,
        qwen36=True,
        source_row=6,
        source_column="910B reuse of row6 910C optimized + dependency script",
        source_cell_refs=("G6", "J6"),
        is_reuse=True,
        notes="910B row is not measured in the source workbook; config reuses 910C model parameters but disables offload by the 910B policy.",
    ),
    _moe(
        name="Qwen3.6-35B-A3B-910B-reuse",
        model_name="Qwen/Qwen3.6-35B-A3B",
        card="910b",
        device_count=2,
        port=7799,
        tp=2,
        max_num_seqs=32,
        max_model_len=131072,
        max_num_batched_tokens=8192,
        gpu_memory_utilization=0.9,
        served_model_name="qwen36",
        mtp_tokens=3,
        qwen36=True,
        source_row=8,
        source_column="910B reuse of row8 910C optimized + dependency script",
        source_cell_refs=("G8", "J6"),
        is_reuse=True,
        notes="910B row is not measured in the source workbook; config reuses 910C model parameters but disables offload by the 910B policy.",
    ),
]


MANAGED_ENV_KEYS = {
    "ASCEND_PLATFORM",
    "AVAILABLE_POD_MEM_SIZE",
    "DEVICE_COUNT",
    "DISTRIBUTED_EXECUTOR_BACKEND",
    "ENABLE_AUTO_THINK_CHOICE",
    "ENABLE_AUTO_TOOL_CHOICE",
    "ENABLE_KV_DISK_OFFLOAD",
    "ENABLE_KV_MEM_OFFLOAD",
    "ENABLE_KV_OFFLOAD",
    "ENABLE_REASON_PROXY",
    "ENABLE_SPARSE",
    "ENABLE_SPECULATIVE_DECODE",
    "ENGINE",
    "ENGINE_IMAGE_FLAVOR",
    "ENGINE_PORT",
    "ENGINE_VERSION",
    "GLOO_SOCKET_IFNAME",
    "HEAD_NODE_ADDR",
    "HOST",
    "INPUT_LENGTH",
    "KV_MEM_OFFLOAD_SIZE",
    "LMCACHE_OFFLOAD",
    "MASTER_IP",
    "MAX_NUM_BATCHED_TOKENS",
    "MAX_NUM_SEQS",
    "MODEL_NAME",
    "MODEL_PATH",
    "NETWORK_INTERFACE",
    "NODE_IPS",
    "NODE_RANK",
    "NODES",
    "NNODES",
    "OUTPUT_LENGTH",
    "POD_IP",
    "PORT",
    "PROXY_PORT",
    "RANK_IP",
    "SAVE_PATH",
    "SD_ENABLE",
    "SPARSE_ENABLE",
    "SPECULATIVE_DECODE_MODEL_PATH",
    "TRUST_REMOTE_CODE",
    "WINGS_ASCEND_PLATFORM",
    "WINGS_HARDWARE_FILE",
}

BOOL_FLAGS = {
    "async-scheduling",
    "enable-auto-tool-choice",
    "enable-expert-parallel",
    "enable-prefix-caching",
    "language-model-only",
    "no-disable-hybrid-kv-cache-manager",
    "no-enable-prefix-caching",
    "trust-remote-code",
}


def _ensure_import_path() -> None:
    sys.path.insert(0, str(WINGS_CONTROL_ROOT))


def _as_posix(path: Path) -> str:
    return str(path).replace("\\", "/")


def _safe_clear_dir(path: Path) -> None:
    root = DRY_RUN_ROOT.resolve()
    target = path.resolve()
    if target == root or root not in target.parents:
        raise RuntimeError(f"refusing to clear path outside dry_run root: {target}")
    path.mkdir(parents=True, exist_ok=True)
    for child in path.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8", newline="\n")


def _reset_env(shared_root: Path) -> None:
    for key in MANAGED_ENV_KEYS:
        os.environ.pop(key, None)
    shared_root.mkdir(parents=True, exist_ok=True)
    os.environ["SHARED_VOLUME_PATH"] = _as_posix(shared_root)
    os.environ["ENABLE_REASON_PROXY"] = "false"
    os.environ["POD_IP"] = "127.0.0.1"
    os.environ["RANK_IP"] = "127.0.0.1"
    os.environ["NETWORK_INTERFACE"] = "lo"
    os.environ["DISTRIBUTED_EXECUTOR_BACKEND"] = "mp"
    os.environ["ENABLE_KV_OFFLOAD"] = "true"
    os.environ["ENABLE_KV_MEM_OFFLOAD"] = "true"
    os.environ["LMCACHE_OFFLOAD"] = "true"
    os.environ["KV_MEM_OFFLOAD_SIZE"] = "40"
    os.environ["AVAILABLE_POD_MEM_SIZE"] = "262144"
    os.environ["ENABLE_KV_DISK_OFFLOAD"] = "false"


def _create_model_dir(root: Path, scenario: Scenario) -> Path:
    model_dir = root / "_model_configs" / scenario.name
    model_dir.mkdir(parents=True, exist_ok=True)
    config = {
        "architectures": [scenario.architecture],
        "model_type": "qwen3_5_moe" if "Moe" in scenario.architecture else "qwen3_5",
        "torch_dtype": "bfloat16",
        "num_hidden_layers": 64,
    }
    if scenario.quantization:
        config["quantization_config"] = {"quant_method": scenario.quantization}
    _write_json(model_dir / "config.json", config)
    _write_text(model_dir / "wings.txt", f"{scenario.model_name}\n")
    _write_json(
        model_dir / "hardware_info.json",
        {
            "device": "ascend",
            "hardware_family": scenario.hardware_family,
        },
    )
    return model_dir


def _source_workbook_audit() -> dict[str, Any]:
    if not EXCEL_STANDARD.exists():
        raise FileNotFoundError(EXCEL_STANDARD)
    workbook = load_workbook(EXCEL_STANDARD, data_only=False)
    worksheet = workbook.active
    model_rows = [
        row
        for row in range(2, worksheet.max_row + 1)
        if str(worksheet.cell(row, 1).value or "").strip().lower().startswith(("qwen/", "eco-tech/"))
    ]
    missing_cells: list[str] = []
    for scenario in STANDARD_SCENARIOS:
        for ref in scenario.source_cell_refs:
            if not str(worksheet[ref].value or "").strip():
                missing_cells.append(f"{scenario.name}:{ref}")
    return {
        "source_file": str(EXCEL_STANDARD.relative_to(REPO_ROOT)),
        "source_sheet": worksheet.title,
        "source_dimensions": f"{worksheet.max_row} rows x {worksheet.max_column} columns",
        "source_model_rows": model_rows,
        "source_model_row_count": len(model_rows),
        "standard_scenario_count": len(STANDARD_SCENARIOS),
        "reuse_scenario_count": len(REUSE_SCENARIOS),
        "missing_source_cells": missing_cells,
        "baseline_workbook_present": EXCEL_BASELINE.exists(),
    }


def _parse_flags(command_line: str) -> dict[str, Any]:
    tokens = shlex.split(command_line, posix=True)
    flags: dict[str, Any] = {}
    i = 0
    while i < len(tokens):
        token = tokens[i]
        if not token.startswith("--"):
            i += 1
            continue
        key = token[2:]
        if key in BOOL_FLAGS or i + 1 >= len(tokens) or tokens[i + 1].startswith("--"):
            flags[key] = True
            i += 1
        else:
            flags[key] = tokens[i + 1]
            i += 2
    return flags


def _extract_engine_command(script: str) -> str:
    for raw in script.splitlines():
        line = raw.strip()
        if "python3 -m vllm.entrypoints.openai.api_server" not in line:
            continue
        if line.startswith("echo "):
            continue
        if line.endswith("&"):
            line = line[:-1].rstrip()
        return line
    raise RuntimeError("unable to locate final vLLM api_server command")


def _json_flag(flags: dict[str, Any], key: str) -> Any:
    value = flags.get(key)
    if value is None or value is True:
        return None
    return json.loads(value)


def _flag_int(flags: dict[str, Any], key: str) -> int | None:
    value = flags.get(key)
    return None if value is None or value is True else int(value)


def _flag_float(flags: dict[str, Any], key: str) -> float | None:
    value = flags.get(key)
    return None if value is None or value is True else float(value)


def _actual_from_command(flags: dict[str, Any], command: str, script: str) -> dict[str, Any]:
    kv_config = _json_flag(flags, "kv-transfer-config")
    return {
        "command": command,
        "port": _flag_int(flags, "port"),
        "tensor_parallel_size": _flag_int(flags, "tensor-parallel-size"),
        "data_parallel_size": _flag_int(flags, "data-parallel-size"),
        "max_num_seqs": _flag_int(flags, "max-num-seqs"),
        "max_model_len": _flag_int(flags, "max-model-len"),
        "max_num_batched_tokens": _flag_int(flags, "max-num-batched-tokens"),
        "gpu_memory_utilization": _flag_float(flags, "gpu-memory-utilization"),
        "seed": _flag_int(flags, "seed"),
        "served_model_name": flags.get("served-model-name"),
        "quantization": flags.get("quantization"),
        "enable_expert_parallel": bool(flags.get("enable-expert-parallel")),
        "trust_remote_code": bool(flags.get("trust-remote-code")),
        "async_scheduling": bool(flags.get("async-scheduling")),
        "language_model_only": bool(flags.get("language-model-only")),
        "enable_auto_tool_choice": bool(flags.get("enable-auto-tool-choice")),
        "tool_call_parser": flags.get("tool-call-parser"),
        "additional_config": _json_flag(flags, "additional-config"),
        "compilation_config": _json_flag(flags, "compilation-config"),
        "speculative_config": _json_flag(flags, "speculative-config"),
        "kv_transfer_config": kv_config,
        "no_disable_hybrid_kv_cache_manager": bool(flags.get("no-disable-hybrid-kv-cache-manager")),
        "has_sparse_config": "sparse-config" in flags or "--sparse-config" in command,
        "has_kv_load_failure_policy": bool(
            isinstance(kv_config, dict) and "kv_load_failure_policy" in kv_config
        ),
        "memcache_meta_present": "WINGS_MEMCACHE_META_SERVICE_URL" in script,
        "memcache_config_present": "WINGS_MEMCACHE_CONFIG_STORE_URL" in script,
    }


def _expected(scenario: Scenario) -> dict[str, Any]:
    spec = {
        "method": "qwen3_5_mtp",
        "num_speculative_tokens": scenario.mtp_tokens,
        "enforce_eager": True,
    }
    return {
        "port": scenario.port,
        "tensor_parallel_size": scenario.tp,
        "data_parallel_size": scenario.dp,
        "max_num_seqs": scenario.max_num_seqs,
        "max_model_len": scenario.max_model_len,
        "max_num_batched_tokens": scenario.max_num_batched_tokens,
        "gpu_memory_utilization": scenario.gpu_memory_utilization,
        "seed": scenario.seed,
        "served_model_name": scenario.served_model_name,
        "quantization": scenario.quantization,
        "enable_expert_parallel": scenario.enable_expert_parallel,
        "trust_remote_code": True,
        "language_model_only": scenario.language_model_only,
        "enable_auto_tool_choice": True,
        "tool_call_parser": "qwen3_coder",
        "additional_config": scenario.additional_config,
        "compilation_config": COMPILATION_CONFIG,
        "speculative_config": spec,
        "kv_transfer_config": QWEN_MEMCACHE_KV_CONFIG if scenario.offload else None,
        "no_disable_hybrid_kv_cache_manager": scenario.offload,
        "has_sparse_config": False,
        "has_kv_load_failure_policy": False,
        "memcache_meta_present": scenario.offload,
        "memcache_config_present": scenario.offload,
    }


def _compare(expected: dict[str, Any], actual: dict[str, Any]) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    for key, exp in expected.items():
        act = actual.get(key)
        if isinstance(exp, float):
            ok = act is not None and abs(float(act) - exp) < 0.000001
        else:
            ok = act == exp
        checks.append({"field": key, "expected": exp, "actual": act, "ok": ok})
    return checks


def _validate_memcache_ports(scenario: Scenario, script: str) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    if scenario.offload:
        assert scenario.memcache_ports is not None
        meta_port, config_port = scenario.memcache_ports
        for label, port in (("memcache_meta_port", meta_port), ("memcache_config_port", config_port)):
            checks.append(
                {
                    "field": label,
                    "expected": port,
                    "actual": port if f"tcp://127.0.0.1:{port}" in script else None,
                    "ok": f"tcp://127.0.0.1:{port}" in script,
                }
            )
        # Dry-run 页面输入固定为节点总容量 40G。MemCache 与 LMCache 使用同一
        # 口径，最终写入 mmc_local.conf 前必须按 device_count 均分。
        expected_dram_gb = max(1, 40 // scenario.device_count)
        dram_export = f'export WINGS_MEMCACHE_DRAM_GB="{expected_dram_gb}"'
        checks.append(
            {
                "field": "memcache_dram_gb_per_card",
                "expected": expected_dram_gb,
                "actual": expected_dram_gb if dram_export in script else None,
                "ok": dram_export in script,
            }
        )
    else:
        for label in ("memcache_meta_port", "memcache_config_port"):
            has_any = "WINGS_MEMCACHE_" in script or "tcp://127.0.0.1:500" in script
            checks.append({"field": label, "expected": None, "actual": "present" if has_any else None, "ok": not has_any})
    return checks


def _build_one(scenario: Scenario, out_root: Path) -> dict[str, Any]:
    from core.port_plan import derive_port_plan
    from core.start_args_compat import parse_launch_args
    from core.wings_entry import build_launcher_plan
    from config.settings import settings

    shared_root = out_root / "_shared"
    _reset_env(shared_root)
    model_dir = _create_model_dir(out_root, scenario)
    os.environ["WINGS_HARDWARE_FILE"] = _as_posix(model_dir / "hardware_info.json")
    os.environ["MODEL_NAME"] = scenario.model_name
    os.environ["MODEL_PATH"] = _as_posix(model_dir)
    os.environ["ENGINE"] = "vllm_ascend"
    os.environ["DEVICE_COUNT"] = str(scenario.device_count)
    os.environ["PORT"] = str(scenario.port)
    os.environ["TRUST_REMOTE_CODE"] = "true"
    os.environ["ENABLE_AUTO_TOOL_CHOICE"] = "true"
    os.environ["ENABLE_SPECULATIVE_DECODE"] = "true"
    os.environ["SPECULATIVE_DECODE_MODEL_PATH"] = "none"
    os.environ["ENABLE_SPARSE"] = "true"
    os.environ["SAVE_PATH"] = _as_posix(out_root / scenario.name)

    argv = [
        "--model-name", scenario.model_name,
        "--model-path", _as_posix(model_dir),
        "--engine", "vllm_ascend",
        "--device-count", str(scenario.device_count),
        "--port", str(scenario.port),
        "--save-path", _as_posix(out_root / scenario.name),
        "--trust-remote-code",
        "--enable-auto-tool-choice",
        "--enable-speculative-decode",
        "--speculative-decode-model-path", "none",
        "--enable-sparse",
        "--node-rank", "0",
    ]
    launch_args = parse_launch_args(argv)
    port_plan = derive_port_plan(
        port=launch_args.port,
        enable_reason_proxy=False,
        health_port=settings.HEALTH_PORT,
    )
    plan = build_launcher_plan(launch_args, port_plan)

    scenario_dir = out_root / scenario.name
    _write_text(scenario_dir / "start_command.sh", plan.command)
    command = _extract_engine_command(plan.command)
    _write_text(scenario_dir / "exec_command.txt", command + "\n")
    flags = _parse_flags(command)
    actual = _actual_from_command(flags, command, plan.command)
    expected = _expected(scenario)
    checks = _compare(expected, actual) + _validate_memcache_ports(scenario, plan.command)
    failures = [check for check in checks if not check["ok"]]

    return {
        "scenario": scenario.name,
        "result": "PASS" if not failures else "FAIL",
        "source": {
            "row": scenario.source_row,
            "column": scenario.source_column,
            "cell_refs": list(scenario.source_cell_refs),
            "reuse": scenario.is_reuse,
            "notes": scenario.notes,
        },
        "hardware_input": {
            "device": "ascend",
            "hardware_family": scenario.hardware_family,
            "device_count": scenario.device_count,
        },
        "expected": expected,
        "actual": {k: v for k, v in actual.items() if k != "command"},
        "checks": checks,
        "failures": failures,
        "start_command": str((scenario_dir / "start_command.sh").relative_to(REPO_ROOT)),
        "exec_command": str((scenario_dir / "exec_command.txt").relative_to(REPO_ROOT)),
        "merged_engine_config": plan.merged_params.get("engine_config", {}),
        "active_smart_features": plan.merged_params.get("_smart_feats", []),
        "allowed_smart_features": plan.merged_params.get("_allowed_smart_feats", []),
    }


def _markdown_report(title: str, audit: dict[str, Any], results: list[dict[str, Any]]) -> str:
    lines = [
        f"# {title}",
        "",
        f"Source workbook: `{audit['source_file']}` (`{audit['source_sheet']}`, {audit['source_dimensions']})",
        f"Model rows: {audit['source_model_row_count']} | standard scenarios: {audit['standard_scenario_count']} | reuse scenarios: {audit['reuse_scenario_count']}",
        "Hardware input standard: minimal hardware_info.json with `device` and `hardware_family`; no `details` input.",
        "Function Call parser expectation: `qwen3_coder` per adaptation decision, even where the source script text still says `hermes`.",
        "Qwen MemCache expectation: AscendStoreConnector config must not contain `kv_load_failure_policy`.",
        "Qwen 910B policy: keep scenario-specific MTP, but suppress offload and all MemCache/LMCache launch fragments.",
        "MemCache memory policy: page memory is node total and is evenly divided by `device_count` before writing per-card DRAM.",
        "",
        "| Scenario | Result | Source | Hardware | Active features | Failed checks | start_command.sh |",
        "|---|---:|---|---|---|---|---|",
    ]
    for result in results:
        failures = result["failures"]
        failed = "-" if not failures else "<br>".join(
            f"{f['field']}: expected `{json.dumps(f['expected'], ensure_ascii=False)}` actual `{json.dumps(f['actual'], ensure_ascii=False)}`"
            for f in failures
        )
        source = f"row{result['source']['row']} / {result['source']['column']}"
        hardware = result["hardware_input"]["hardware_family"]
        active = ",".join(result["active_smart_features"]) or "-"
        lines.append(
            f"| {result['scenario']} | {result['result']} | {source} | `{hardware}` | `{active}` | {failed} | `{result['start_command']}` |"
        )
    lines.append("")
    lines.append("## Checked Fields")
    lines.append("")
    lines.append(
        "`TP`, `DP`, `port`, `served_model_name`, max sequence/token limits, GPU memory utilization, seed, "
        "`tool_call_parser`, Function Call switch, MTP method/token/enforce_eager, EP, quantization, "
        "`additional_config`, `compilation_config`, `language_model_only`, sparse absence, MemCache ports, "
        "and absence of Qwen `kv_load_failure_policy`."
    )
    lines.append("")
    if any(r["failures"] for r in results):
        lines.append("Overall result: FAIL")
    else:
        lines.append("Overall result: PASS")
    lines.append("")
    return "\n".join(lines)


def _run_group(title: str, out_root: Path, scenarios: list[Scenario], audit: dict[str, Any]) -> list[dict[str, Any]]:
    _safe_clear_dir(out_root)
    results = [_build_one(scenario, out_root) for scenario in scenarios]
    payload = {"audit": audit, "results": results}
    _write_json(out_root / "comparison.json", payload)
    _write_text(out_root / "comparison.md", _markdown_report(title, audit, results))
    return results


def main() -> int:
    _safe_clear_dir(STANDARD_OUT_ROOT)
    _safe_clear_dir(REUSE_OUT_ROOT)
    os.environ.setdefault("SHARED_VOLUME_PATH", _as_posix(STANDARD_OUT_ROOT / "_shared"))
    _ensure_import_path()

    audit = _source_workbook_audit()
    if audit["source_model_row_count"] != 8:
        raise RuntimeError(f"unexpected source model row count: {audit['source_model_row_count']}")
    if audit["standard_scenario_count"] != 11:
        raise RuntimeError(f"unexpected standard scenario count: {audit['standard_scenario_count']}")
    if audit["missing_source_cells"]:
        raise RuntimeError(f"source workbook missing cells: {audit['missing_source_cells']}")

    standard_results = _run_group(
        "Qwen Day0 current dry-run comparison",
        STANDARD_OUT_ROOT,
        STANDARD_SCENARIOS,
        audit,
    )
    reuse_results = _run_group(
        "Qwen Day0 910B reuse dry-run comparison",
        REUSE_OUT_ROOT,
        REUSE_SCENARIOS,
        audit,
    )
    all_results = standard_results + reuse_results
    failed = [r for r in all_results if r["failures"]]
    print(f"standard={len(standard_results)} reuse={len(reuse_results)} failed={len(failed)}")
    print(STANDARD_OUT_ROOT / "comparison.md")
    print(REUSE_OUT_ROOT / "comparison.md")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
